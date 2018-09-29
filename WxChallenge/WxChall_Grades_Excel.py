import os;
import numpy as np;
from openpyxl import Workbook;
from openpyxl.utils import get_column_letter as col2letter;
from pandas import read_csv;
from WxChallenge import WxChallenge;
from WxChallenge.utils import fix_Roster_CSV;
from WxChallenge.data import fcst_tag, fname_tag, lname_tag, class_tag, grd_df_cols;

################################################################################
class ExcelBook( Workbook ):
  '''
  A class for generating an XLSX file for a given class that
  forecasters are enrolled in.
  '''
  def __init__(self, file):
    Workbook.__init__(self);                                                    # Initialize Workbook
    self.remove(self.active);                                                   # Remove the initial sheet
    self.file   = file;                                                         # File to save to on saveBook method
    self.fcstrs = [];                                                           # List to store (forecaster id, last name, first name) tuples for the Workbook
    self._rows  = {};                                                           # Dictionary to store the current row for a given sheet
  ##############################################################################
  def fcstrExist(self, fcstrName):
    '''
    Name:
       fcstrExist
    Purpose:
       Method to check if given forecaster should be added to the Workbook.
    Inputs:
       fcstrName : Name of the forecaster
    Outputs:
       Returns True if forecaster should be added.
       Returns None if forecaster should NOT be added.
    Keywords:
       None.
    '''
    for fcstr in self.fcstrs:                                                   # Iterate over all tuples in the fcstrs list
      if fcstrName == fcstr[0]: return list(fcstr);                             # If the fcstrName input matches the forecaster id in the tuple, then return the tuple
    return None;                                                                # If made it here, then did NOT find the forecaster, so return None
  ##############################################################################
  def addSheet(self, sheetName):
    '''
    Name:
       addSheet
    Purpose:
       Method to add a sheet to the Workbook given a sheet Name
    Inputs:
       sheetName : Name of the sheet to add
    Outputs:
       None.
    Keywords:
       None.
    '''
    if sheetName not in self:                                                   # If a sheet with name sheetName does NOT exist
      self.create_sheet( sheetName );                                           # Create a new sheet
      self._rows[sheetName] = 1;                                                # Set the row counter for the sheet to 1
      columns = ['last name', 'first name'] + grd_df_cols;                      # Set column names for the sheet
      for i in range( len(columns) ):                                           # Iterate over the columns for the sheet
        self[sheetName].cell( self._rows[sheetName], i+1, columns[i] );         # Write each column to the sheet
      self._rows[sheetName] += 1;                                               # Increment the row number of the sheet
  ##############################################################################
  def addData(self, sheetName, fcstr, values):
    '''
    Name:
       addData
    Purpose:
       Method to add a data to a given sheet in the Workbook
    Inputs:
       sheetName : Name of the sheet to add data to
       fcstr     : Tuple of (forecaster id, last name, first name)
       values    : Values to add to the Sheet; i.e., a row
    Outputs:
       None.
    Keywords:
       None.
    '''
    values = fcstr[1:] + values.tolist();                                       # Set list of values to write to the sheet
    for col in range( len(values) ):                                            # Iterate over all values in values list
      self[sheetName].cell( self._rows[sheetName], col+1, values[col] );        # Write the data to the SpreadSheet
    self._rows[sheetName] += 1;                                                 # Increment the row by 1
  ##############################################################################
  def saveBook(self):
    '''
    Name:
       saveBook
    Purpose:
       Method to save the Workbook to a file
    Inputs:
       None.
    Outputs:
       None.
    Keywords:
       None.
    '''
    self.__addFinalScore();                                                     # Add the final Scores sheet
    self.__addSorting();                                                        # Enable sorting on all sheets
    dir = os.path.dirname( self.file );                                         # Parent directory of file path
    if not os.path.isdir( dir ): os.makedirs( dir );                            # If the output directory does NOT exist, create it
    self.save( self.file );                                                     # Save the Workbook object to a file
  ##############################################################################
  def __addSorting(self):
    '''
    Name:
       addSorting
    Purpose:
       Method to add sorting of columns the sheets in the Workbook
    Inputs:
       None.
    Outputs:
       None.
    Keywords:
       None.
    '''
    fmt    = "{}{}:{}{}";                                                       # Format for cell ranges
    for sheet in self:                                                          # Iterate over all sheets in the Workbook
      rows = [ 1,   self._rows[sheet.title] ];                                  # Set rows from second through last row
      cols = [ 'A', col2letter(sheet.max_column) ];                             # Set columns to first through last column
      sheet.auto_filter.ref = fmt.format(cols[0], rows[0], cols[1], rows[1]);   # Set the reference space for the filtering; i.e., which data will be filtered
      for i in range( sheet.max_column ):                                       # Iterate ver all the columns
        col = col2letter( i+1 );                                                # Get the leter name for the column
        rng = fmt.format( col, rows[0], col, rows[1] );                         # Set up the range of values for sorting of single column
        sheet.auto_filter.add_sort_condition( rng );                            # Define sorting for the column
  ##############################################################################
  def __addFinalScore(self):
    '''
    Name:
       addFinalScore
    Purpose:
       Method to add a sheet containing final scores that are the
       average of all scores in previous sheets.
    Inputs:
       None.
    Outputs:
       None.
    Keywords:
       None.
    '''
    scores  = [];                                                               # List to store scores in
    for sheet in self:                                                          # Iterate over all sheets in the Workbook
      col = col2letter( sheet.max_column );                                     # Convert the maximum column number to letter
      scores.append( [ i.value for i in sheet[col][1:] ] );                     # Get the values for the last column, excluding the header
    scores = np.asarray(scores).mean( axis = 0 );                               # Covert the scores list to a numpy array and then average over the zeroth axis
    names  = sheet['A1':'B{}'.format(self.active.max_row)];                     # Get the first and last names (first two columns) from the last sheet of the for loop
    ws     = self.create_sheet('Final_Grades');                                 # Create new sheet title Final_Grades
    for i in range( len(names) ):                                               # Iterate over all the tuples in names
      row = [ j.value for j in names[i] ];                                      # Get the value in the tuple and convert to list
      row.append( 'Grade' if i == 0 else scores[i-1] );                         # Append a header, or final grade, to the row list depending on value of i
      ws.append(row);                                                           # Write the row the new sheet
    self._rows[ws.title] = ws.max_row;                                          # Set the current row value in the _rows attribute
  
################################################################################
class WxChall_Grades_Excel( object ):
  '''
  A class for computing grades and output to Excel SpreadSheets.
  A SpreadSheet will be created for each class that is defined
  in the roster. In each class SpreadSheet there will be a
  sheet for each forecast city.
  '''
  def __init__(self, semester, year, roster, school=None, verbose=False, outdir=None):
    '''
    Name:
       __init___
    Purpose:
       To initialize the WxChall_Grades_Excel class
    Inputs:
       semester  : String of the semester of the data; 'fall' or 'spring'
       year      : Integer year
       roster    : Path to the roster CSV for the school; local Wx manager
                    has access to this.
    Outputs:
       Excel SpreadSheet(s)
    Keywords:
       school   : 3-character school tag
       verbose  : Set to increase verbosity
       outdir   : Top level output directory for SpreadSheet files.
                    Default is same location as roster file
    '''
    inst  = WxChallenge();                                                      # Initialize the WxChallenge
    self.fcsts = inst.get_forecasts( 
      school   = school,
      semester = semester,
      year     = year
    );                                                                          # Get forecasts based on command line arguments
    inst.close();                                                               # Close SQL database
    if self.fcsts is None:                                                      # If no forecasts returned
      print( 'No forecasts found!' );                                           # Print a message
      return False;                                                             # Exit
    self.outdir = os.path.dirname( roster ) if outdir is None else outdir;      # Set output directory based on roster CSV path if no outdir specified
    self.verbose = verbose;
    self.Workbooks = self.initWorkbooks( roster );                              # Parse data from roster CSV
  ##############################################################################
  def grades(self):
    '''
    Name:
       grades
    Purpose:
       Main method for computing grades and saving data
       to Excel SpreadSheets.
    Inputs:
       None.
    Outputs:
       Will create Excel SpreadSheets
    Keywords:
       None.
    '''
    consensus, climo = self.getConsensClimo();                                  # Get consensus and climatology data
    for f in self.fcsts:                                                        # Iterate over all the forecasts again
      g = f.calc_grade(
        climatology    = climo,
        sch_consensus  = consensus[f.school],
        ntnl_consensus = consensus['xxx'],
        verbose        = self.verbose
      );                                                                        # Calculate the grade for a forecaster  
      self.updateSpreadSheets( f );                                             # Call method to update the spreadsheets with the current forecaster's grades
    self.saveSpreadSheets();                                                    # Save all the spreadsheets
  ##############################################################################
  def getConsensClimo(self):
    '''
    Name:
       getConsensClimo
    Purpose:
       A method to locate and return consensus and climatology data
       from the list of forecasts
    Inputs:
       None.
    Outputs:
       Returns a dictionary containing consensus data and
       a list containing climatology data
    Keywords:
       None.
    '''
    consensus, climo = {}, [];                                                  # Initialize a dictionary for consensus (school & national) and a list for climatology
    for f in self.fcsts:                                                        # Iterate over all forecasters in list
      if f.is_climo:                                                            # If a forecaster is climo
        if not any( [c == f for c in climo] ): climo.append(f);                 # If the forecaster is NOT in the list, then append it
      elif f.is_consen and f.school not in consensus:                           # Else, if the forecaster is consensus
        consensus[f.school] = f;                                                # Add it to the dictionary
    return consensus, climo;
  ##############################################################################
  def saveSpreadSheets(self):
    '''
    Name:
       saveSpreadSheets
    Purpose:
       A method to save all the Workbook objects to
       Excel SpreadSheet files.
    Inputs:
       None.
    Outputs:
       Generate output files
    Keywords:
       None.
    '''
    for key in self.Workbooks: self.Workbooks[key].saveBook();                  # Iterate over all the keys in the Workbooks dictionary and save the books
  ##############################################################################
  def updateSpreadSheets(self, fcstr):
    '''
    Name:
       updateSpreadSheets
    Purpose:
       A method to add a single forecasters grades to a spreadsheets.
       SpreadSheets are created for each class in the roster CSV
    Inputs:
       fcstr : A WxForecaster object that has grades calculated
    Outputs:
       None.
    Keywords:
       None.
    '''
    if len(self.Workbooks) == 0: return None;                                   # If there is no Workbooks data, return
    for cls in self.Workbooks:                                                  # Iterate over all Workbooks in the Workbooks dictionary
      info = self.Workbooks[cls].fcstrExist( fcstr.name );                      # Check if the forecaster should be in the given Workbook
      if info:                                                                  # If the forecaster is in one of the Workbooks
        for city in fcstr.grades.index:                                         # Iterate over all cities (rows) of the forecaster's grades
          self.Workbooks[cls].addSheet( city );                                 # Add new sheet to the Excel book if one does NOT exist
          self.Workbooks[cls].addData( city, info, fcstr.grades.loc[city] );    # Add data to the sheet
  ##############################################################################
  def initWorkbooks( self, csvfile ):
    '''
    Name:
       initWorkbooks
    Purpose:
       A method to parse data from the roster CSV and generate
       a ExcelBook object for each class that forecasters are
       enrolled in.
    Inputs:
       csvfile  : Full path to the roster CSV file
    Ouputs:
       Returns a dictionary of ExcelBook objects
    Keywords:
       None.
    '''
    Workbooks = {}
    if not os.path.isfile( csvfile ):                                           # If the file does NOT exists
      print( 'Roster file NOT found!' );                                        # Print a message
    else:                                                                       # Else, it exists
      fix_Roster_CSV( csvfile );                                                # Fix the CSV file
      roster = read_csv( csvfile );                                             # Read in the data
      for col in class_tag:                                                     # Iterate over the class ids
        for cls in roster[col].unique():                                        # Iterate over the unique values in the column
          if 'NO CLASS' in cls.upper(): continue;                               # Skip the 'No Class' class
          if cls not in Workbooks:                                              # If the cls key is NOT in the Workbooks dictionary,
            xlsFile = '{}.xlsx'.format(  '_'.join( cls.split() ) );             # Base name for the Excel file
            xlsFile = os.path.join( self.outdir, xlsFile );                     # Full path for the file
            Workbooks[cls] = ExcelBook( xlsFile );                              # Initialize ExcelBook object
      for col in class_tag:                                                     # Iterate over the columns to sort students into Workbooks
        for cls in Workbooks:                                                   # Iterate over all tags in the Workbooks dictionary
          tmp = roster.loc[ roster[col] == cls ];                               # Locate all rows where the class matches
          if len(tmp) > 0:                                                      # If rows located
            fcstID = tmp[fcst_tag].values;                                      # Get forecaster id values
            lname  = tmp[lname_tag].values;                                     # Get forecaster last name values
            fname  = tmp[fname_tag].values;                                     # Get forecaster first name values
            Workbooks[cls].fcstrs += zip( fcstID, lname, fname );               # Zip up the forecaster id, last name, and first name and append values to the 'fcstrs' attribute of the ExcelBook object
    return Workbooks;                                                           # Return the Workbooks dictionary