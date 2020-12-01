import logging
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as col2letter
from pandas import read_csv
from WxChallenge.WxChallenge import WxChallenge

from .roster import fix_Roster_CSV
from .data import fcst_tag, fname_tag, lname_tag, class_tag, grd_df_cols


################################################################################
class ExcelBook( Workbook ):
  """
  For generating XLSX file for a given class that forecasters are enrolled in
  """

  def __init__(self, file):
    Workbook.__init__(self)                                                     # Initialize Workbook
    self.remove(self.active)                                                    # Remove the initial sheet
    self.file   = file                                                          # File to save to on saveBook method
    self.fcstrs = []                                                            # List to store (forecaster id, last name, first name) tuples for the Workbook
    self._rows  = {}                                                            # Dictionary to store the current row for a given sheet

  def fcstrExist(self, fcstrName):
    """
    Check if given forecaster should be added to the Workbook.

    Arguments:
      fcstrName (str): Name of the forecaster

    Keyword arumgent:
      None.

    Returns:
      bool: True for forecaster should be added, False if already exist
    """
    for fcstr in self.fcstrs:                                                   # Iterate over all tuples in the fcstrs list
      if fcstrName == fcstr[0]: return list(fcstr)                              # If the fcstrName input matches the forecaster id in the tuple, then return the tuple
    return None                                                                 # If made it here, then did NOT find the forecaster, so return None

  def addSheet(self, sheetName):
    """
    Add a sheet to the Workbook given a sheet Name

    Arguments:
      sheetName (str): Name of the sheet to add

    Keyword arguments:
       None.

    Returns:
       None.

    """
    if sheetName not in self:                                                   # If a sheet with name sheetName does NOT exist
      self.create_sheet( sheetName )                                            # Create a new sheet
      self._rows[sheetName] = 1                                                 # Set the row counter for the sheet to 1
      columns = ['last name', 'first name'] + grd_df_cols                       # Set column names for the sheet
      for i in range( len(columns) ):                                           # Iterate over the columns for the sheet
        self[sheetName].cell( self._rows[sheetName], i+1, columns[i] )          # Write each column to the sheet
      self._rows[sheetName] += 1                                                # Increment the row number of the sheet

  def addData(self, sheetName, fcstr, values):
    """
    Add a data to a given sheet in the Workbook

    Arguments:
       sheetName (str)  : Name of the sheet to add data to
       fcstr     (tuple): forecaster id, last name, first name)
       values    (iter) : Values to add to the Sheet; i.e., a row

    Keyword arguments:
       None.

    Returns:
       None.

    """
 
    self.addSheet( sheetName ) 
    values = fcstr[1:] + values.tolist()                                        # Set list of values to write to the sheet
    for col in range( len(values) ):                                            # Iterate over all values in values list
      self[sheetName].cell( self._rows[sheetName], col+1, values[col] )         # Write the data to the SpreadSheet
    self._rows[sheetName] += 1                                                  # Increment the row by 1

  def saveBook(self):
    """
    Save the Workbook to a file

    Arguments:
      None.

    Keyword arguments:
      None.

    Returns:
      None.

    """

    self.__addFinalScore()                                                      # Add the final Scores sheet
    self.__addSorting()                                                         # Enable sorting on all sheets
    dir = os.path.dirname( self.file )                                          # Parent directory of file path
    if not os.path.isdir( dir ): os.makedirs( dir )                             # If the output directory does NOT exist, create it
    self.save( self.file )                                                      # Save the Workbook object to a file

  def __addSorting(self):
    """
    Add sorting of columns the sheets in the Workbook

    Arguments:
      None.

    Keyword arguments:
      None.

    Returns:
      None.

    """

    fmt    = "{}{}:{}{}"                                                        # Format for cell ranges
    for sheet in self:                                                          # Iterate over all sheets in the Workbook
      rows = [ 1,   self._rows[sheet.title] ]                                   # Set rows from second through last row
      cols = [ 'A', col2letter(sheet.max_column) ]                              # Set columns to first through last column
      sheet.auto_filter.ref = fmt.format(cols[0], rows[0], cols[1], rows[1])    # Set the reference space for the filtering; i.e., which data will be filtered
      for i in range( sheet.max_column ):                                       # Iterate ver all the columns
        col = col2letter( i+1 )                                                 # Get the leter name for the column
        rng = fmt.format( col, rows[0], col, rows[1] )                          # Set up the range of values for sorting of single column
        sheet.auto_filter.add_sort_condition( rng )                             # Define sorting for the column

  def __addFinalScore(self):
    """
    Add sheet containing final scores 
    
    The final scores are the average of all scores in previous sheets.

    Arguments:
      None.

    Keyword arguments:
      None.

    Returns:
      None.

    """

    sch_bonus = [] 
    ntl_bonus = [] 
    scores    = []                                                              # List to store scores in
    for sheet in self:                                                          # Iterate over all sheets in the Workbook
      nCols = sheet.max_column 
      sch_col = col2letter( nCols - 2 ) 
      ntl_col = col2letter( nCols - 1 ) 
      scr_col = col2letter( nCols     )                                         # Convert the maximum column number to letter

      sch_bonus.append( [ i.value for i in sheet[sch_col][1:] ] ) 
      ntl_bonus.append( [ i.value for i in sheet[ntl_col][1:] ] ) 
      scores.append(    [ i.value for i in sheet[scr_col][1:] ] )               # Get the values for the last column, excluding the header

    sch_bonus = np.asarray(sch_bonus).sum(  axis = 0 )  
    ntl_bonus = np.asarray(ntl_bonus).sum(  axis = 0 )  
    scores    = np.asarray(scores   ).mean( axis = 0 )                          # Covert the scores list to a numpy array and then average over the zeroth axis

    names     = sheet['A1':'B{}'.format(self.active.max_row)]                   # Get the first and last names (first two columns) from the last sheet of the for loop
    ws        = self.create_sheet('Final_Grades')                               # Create new sheet title Final_Grades
    for i in range( len(names) ):                                               # Iterate over all the tuples in names
      row = [ j.value for j in names[i] ]                                       # Get the value in the tuple and convert to list
      if i == 0:
        row = row + ['Consen. School', 'Consen. Ntnl.', 'Grade'] 
      else:
        row = row + [sch_bonus[i-1], ntl_bonus[i-1], scores[i-1]]
      ws.append(row)                                                            # Write the row the new sheet
    self._rows[ws.title] = ws.max_row                                           # Set the current row value in the _rows attribute
  

class WxChall_Grades_Excel( object ):
  """
  For computing grades and output to Excel SpreadSheets

  A SpreadSheet will be created for each class that is defined
  in the roster. In each class SpreadSheet there will be a
  sheet for each forecast city.
  """

  def __init__(self, semester, year, roster, school=None, verbose=False):
    """
    To initialize the WxChall_Grades_Excel class

    Arguments:
      semester  (str): Semester of the data; 'fall' or 'spring'
      year      (int): year
      roster    (str): Path to the roster CSV for the school; local Wx manager
        has access to this.

    Keyword arguments:
      school   (str): 3-character school tag
      verbose  (bool): Set to increase verbosity
      outdir   (str): Top level output directory for SpreadSheet files.
        Default is same location as roster file

    Returns:
      Excel SpreadSheet(s)

    """

    self.__log = logging.getLogger(__name__)
    self.wx  = WxChallenge()                                                    # Initialize the WxChallenge
    self.semester  = semester
    self.year      = year
    self.roster    = roster
    self.school    = school
    self.verbose   = verbose
    self.outdir    = None
    self.Workbooks = None                                                       # Parse data from roster CSV

  def grades(self, outdir = None, vacation = None):
    """
    Compute grades and save data to Excel SpreadSheets

    Arguments:
      None.

    Keyword arguments:
      outdir (str) : Set output directory for Excel files
      vacation (list) : List of tuples wtih start and end dates of any vacations

    Returns:
      Will create Excel SpreadSheets

    """
    
    self.outdir    = os.path.dirname(self.roster) if outdir is None else outdir
    self.Workbooks = self.initWorkbooks( self.roster )
    fcsts = self.wx.get_forecasts( 
      school   = self.school,
      semester = self.semester,
      year     = self.year
    )                                                                           # Get forecasts based on command line arguments

    if len(fcsts) == 0:                                                         # If no forecasts returned
      self.__log.error( 'No forecasts found!' )                                 # Print a message
      return False                                                              # Exit
    model = self.wx.get_forecasts( 
      school   = self.school,
      semester = self.semester,
      year     = self.year,
      models   = True
    )                                                                           # Get forecasts based on command line arguments
    verify = set( fcsts.date.values )
    fcsts.calc_grades( model, verify = self.wx.get_verification( verify ), vacation = vacation )
    for f in fcsts.iterForecasters(grades = True):                              # Iterate over all the forecasts again
      self.updateSpreadSheets( f )                                              # Call method to update the spreadsheets with the current forecaster's grades
    self.saveSpreadSheets()                                                     # Save all the spreadsheets

  def getConsensClimo(self):
    """
    Locate and return consensus and climatology data from list of forecasts

    Arguments:
      None.

    Keyword arguments:
      None.

    Returns:
      (tuple) : Dictionary containing consensus data and a list containing
        climatology data

    """

    consensus, climo = {}, []                                                   # Initialize a dictionary for consensus (school & national) and a list for climatology
    for f in self.fcsts:                                                        # Iterate over all forecasters in list
      if f.is_climo:                                                            # If a forecaster is climo
        if not any( [c == f for c in climo] ): climo.append(f)                  # If the forecaster is NOT in the list, then append it
      elif f.is_consen and f.school not in consensus:                           # Else, if the forecaster is consensus
        consensus[f.school] = f                                                 # Add it to the dictionary
    return consensus, climo

  def saveSpreadSheets(self):
    """
    Save all the Workbook objects to Excel SpreadSheet files

    Arguments:
       None.

    Keyword arguments:
       None.

    Returns:
       Generate output files

    """

    for key in self.Workbooks: self.Workbooks[key].saveBook()                   # Iterate over all the keys in the Workbooks dictionary and save the books

  def updateSpreadSheets(self, fcstr):
    """
    Add a single forecaster's grades to a spreadsheets

    As SpreadSheets are created for each class in the roster CSV, this will
    add a forecaster to the correst sheet based on the classes they
    are in

    Arguments:
      fcstr (WxForecaster): WxForecaster object that has grades calculated

    Keyword arguments:
      None.

    Returns:
      None.

    """

    if len(self.Workbooks) == 0: return None                                    # If there is no Workbooks data, return
    for cls in self.Workbooks:                                                  # Iterate over all Workbooks in the Workbooks dictionary
      info = self.Workbooks[cls].fcstrExist( fcstr.name )                       # Check if the forecaster should be in the given Workbook
      if info:                                                                  # If the forecaster is in one of the Workbooks
        cities = fcstr.getCities()                                              # Get list of all cities for the forecaster
        for city in cities:                                                     # Iterate over all cities (rows) of the forecaster's grades
          values = fcstr.loc[city == cities].values                             # Get values for data in city
          if values.shape[0] == 1:
            values = values.flatten( )
            self.Workbooks[cls].addData( city, info, values )                   # Add data to the sheet

  def initWorkbooks( self, csvfile ):
    """
    Parse data from the roster CSV to generate ExcelBook objects
    
    The data from the roster CSV is parsed and an ExcelBook object
    is generated for each class that forecasters are enrolled in.

    Arguments:
       csvfile  : Full path to the roster CSV file

    Keyword arguments:
       None.

    Returns:
       Returns a dictionary of ExcelBook objects
    
    """

    Workbooks = {}
    if not os.path.isfile( csvfile ):                                           # If the file does NOT exists
      self.__log.error( 'Roster file NOT found!' )                              # Print a message
    else:                                                                       # Else, it exists
      fix_Roster_CSV( csvfile )                                                 # Fix the CSV file
      roster = read_csv( csvfile )                                              # Read in the data
      for col in class_tag:                                                     # Iterate over the class ids
        for cls in roster[col].unique():                                        # Iterate over the unique values in the column
          if 'NO CLASS' in cls.upper(): continue                                # Skip the 'No Class' class
          if cls not in Workbooks:                                              # If the cls key is NOT in the Workbooks dictionary,
            xlsFile = '{}.xlsx'.format(  '_'.join( cls.split() ) )              # Base name for the Excel file
            xlsFile = os.path.join( self.outdir, xlsFile )                      # Full path for the file
            Workbooks[cls] = ExcelBook( xlsFile )                               # Initialize ExcelBook object
      for col in class_tag:                                                     # Iterate over the columns to sort students into Workbooks
        for cls in Workbooks:                                                   # Iterate over all tags in the Workbooks dictionary
          tmp = roster.loc[ roster[col] == cls ]                                # Locate all rows where the class matches
          if len(tmp) > 0:                                                      # If rows located
            fcstID = tmp[fcst_tag].values                                       # Get forecaster id values
            lname  = tmp[lname_tag].values                                      # Get forecaster last name values
            fname  = tmp[fname_tag].values                                      # Get forecaster first name values
            Workbooks[cls].fcstrs += zip( fcstID, lname, fname )                # Zip up the forecaster id, last name, and first name and append values to the 'fcstrs' attribute of the ExcelBook object
    return Workbooks                                                            # Return the Workbooks dictionary
